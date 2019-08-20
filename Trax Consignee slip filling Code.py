from tkinter import Button, Label, Frame, Tk, StringVar, Entry
from openpyxl import Workbook, load_workbook
from tkinter.filedialog import askdirectory, askopenfilename
import os
from datetime import datetime

date = datetime.now()

service_type_id = '1'
pickup_address_id = "2203"
item_product_type_id = '24'
item_quantity = '1'
special_instructions = "Handle with care, call before delivery"
estimated_weight = '0.1'
mode_of_shipment_id = {'overnight':'1','overland':'2'}
mode_of_payment_id = '1'
charges_mode_id = '4'
pickup_Date = date.strftime('%Y-%m-%d')

file_get_position = ''
file_set_position = ''

class Gui(Tk):
    def __init__(self):
        super().__init__()
    def window_size(self,width=700,height=500):
        self.geometry(f'{width}x{height}')


def str_reverse(st):
    st = list(st)
    st.reverse()
    return ''.join(st)

def get_file(var):
    gui.file_get_dir = askopenfilename(filetype=(('Excel file', '*.xlsx'),('All files', '*.*')))
    var.set('')
    var.set(gui.file_get_dir)

def file_destination(var):
    file_set_position = askdirectory()
    var.set('')
    var.set(file_set_position)
def done():

    try:
        st = str_reverse(e1.get())
        index = st.index('/')
        file_xlsx = e1.get()[-index:]
        file_get_position = e1.get()[:-index]

        os.chdir(file_get_position)

        # a checker is available for correct format of file
        # an exception handler is for checking that the file is present or not
        if ".xlsx" not in file_xlsx:
            error.set('Incorrect file format or type, try again')
        else:
            error.set('')
            load_wb = load_workbook(file_xlsx)  # creat an object of Load_workbook and get data from an xl file
            sheet_list = list(load_wb.sheetnames)  # get list of number of sheets and store it in a variable
            load_wb_sheet = load_wb[sheet_list[0]]  # get the load workbook sheet and store it in a variable

            heading_list = ['Service Type ID', 'Pickup Address ID', 'Show Information on Air Waybill',
                            'Consignee City Name',
                            'Consignee Name', 'Consignee Address', 'Consignee Phone Number 1 (03000000000)',
                            'Consignee Phone Number 2 (03000000000)', 'Consignee Email Address', 'Order ID',
                            'Item Product Type ID', 'Item Description', 'Item Quantity', 'Item Insurance',
                            'Product Value',
                            'Replacement Item Product Type ID', 'Replacement Item Description',
                            'Replacement Item Quantity',
                            'Pickup Date (YYYY-MM-DD)', 'Special Instructions', 'Estimated Weight (kg)',
                            'Mode of Shipment ID',
                            'Same Day Timing ID', 'Collection Amount', 'Mode of Payment ID', 'Charges Mode ID']
            creat_wb = Workbook()  # creat an object of Workwook class
            creat_wb_sheet = creat_wb['Sheet']  # get the xl sheet and by name and store it in a variable
            num = 0
            for x in heading_list:  # this loop will initialize the creat_wb_sheet with items of heading list
                num += 1
                creat_wb_sheet.cell(row=1, column=num).value = x

            row = 1
            for x in range(1, load_wb_sheet.max_row + 1):  # this loop will run for max rows in load_wb_sheet times
                row += 1
                for y in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:  # number of columns in load_wb_sheet

                    if y == 'A':
                        creat_wb_sheet[f'A{row}'].value = service_type_id
                        creat_wb_sheet[f'B{row}'].value = pickup_address_id
                        creat_wb_sheet[f'C{row}'].value = 'no'
                        creat_wb_sheet[f'J{row}'].value = load_wb_sheet[f'A{x}'].value

                    elif y == 'B':
                        creat_wb_sheet[f'K{row}'].value = item_product_type_id

                    elif y == 'C':
                        creat_wb_sheet[f'E{row}'].value = load_wb_sheet[f'C{x}'].value

                    elif y == 'D':
                        creat_wb_sheet[f'L{row}'].value = load_wb_sheet[f'D{x}'].value
                        creat_wb_sheet[f'M{row}'].value = item_quantity
                        creat_wb_sheet[f'N{row}'].value = 'no'

                    elif y == 'E':
                        creat_wb_sheet[f'S{row}'].value = pickup_Date
                        creat_wb_sheet[f'T{row}'].value = special_instructions
                        creat_wb_sheet[f'U{row}'].value = estimated_weight

                    elif y == 'F':
                        creat_wb_sheet[f'V{row}'].value = mode_of_shipment_id['overnight']
                        creat_wb_sheet[f'X{row}'].value = load_wb_sheet[f'F{x}'].value

                    elif y == 'G':
                        creat_wb_sheet[f'Y{row}'].value = mode_of_payment_id
                        creat_wb_sheet[f'Z{row}'].value = charges_mode_id

                    elif y == 'H':
                        creat_wb_sheet[f'F{row}'].value = load_wb_sheet[f'H{x}'].value

                    elif y == 'I':
                        creat_wb_sheet[f'D{row}'].value = load_wb_sheet[f'I{x}'].value

                    elif y == 'J':
                        creat_wb_sheet[f'G{row}'].value = load_wb_sheet[f'J{x}'].value

                    else:
                        pass

            load_wb.close()

            os.chdir(e2.get())
            creat_wb.save('Shipment Booking list.xlsx')

    except Exception as f:
        error.set(f)

gui = Gui()
gui.window_size()
gui.title('Idealancy COD Slip')
gui.config(bg= 'white')
Label(text='IDEALANCY',font='arial 20 bold',bg='white').pack(fill='x')

get_file_entry = StringVar()
f1 = Frame(bg='White').pack(side='left')
Label(f1,text='Get File',font='arial 13 bold',bg='White').pack(anchor='w',side='top')
Button(f1,text="Browse File",command=lambda:get_file(get_file_entry)).pack(anchor='w')
e1 = Entry(f1,textvariable=get_file_entry,border='3')
e1.pack(anchor='w',ipadx=150)

Label(text='__________________________________',bg='white').pack(anchor='w',pady=30)

file_destination_entry = StringVar()
f2 = Frame(bg='White',pady=30).pack(side='left')
Label(f2,text='File Destinaton',font='arial 13 bold',bg='White').pack(anchor='w',side='top')
Button(f2,text="Browse File",command=lambda:file_destination(file_destination_entry)).pack(anchor='w')
e2 = Entry(f2,textvariable=file_destination_entry,border='3')
e2.pack(anchor='w',ipadx=150)

error = StringVar()

Button(text='Done',command=done).pack(side='bottom',anchor='e',ipadx=30,ipady=30)

Label(textvariable=error,fg='red',bg='White',font='2').pack(side='top',anchor='w',pady=30)

gui.mainloop()

