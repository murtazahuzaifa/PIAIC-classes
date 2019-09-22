from tkinter import Button, Label, Frame, Tk, StringVar, Entry
from openpyxl import Workbook, load_workbook
from tkinter.filedialog import askdirectory, askopenfilename
import os
from datetime import datetime

date = datetime.now()

service_type_id = '1'
pickup_address_id = "2203"
item_product_type_id = '24'
item_quantity = '1'
special_instructions = "Handle with care, call before delivery"
estimated_weight = '0.1'
mode_of_shipment_id = {'overnight':'1','overland':'2'}
mode_of_payment_id = '1'
charges_mode_id = '4'
pickup_Date = date.strftime('%Y-%m-%d')

file_get_position = ''
file_set_position = ''

list_of_cities = ['4PL',
 'Abbottabad',
 'Abdul Hakim',
 'Ahmed Pur East',
 'Akora Khattak',
 'Ali Pur Chatta',
 'Alipur',
 'Ambar',
 'Arifwala',
 'Attock',
 'Badin',
 'Bagh',
 'Bahawalnagar',
 'Bahawalpur',
 'Bala Pir',
 'Bannu',
 'Basir Pur',
 'Basti Lar',
 'Basti Malook',
 'Basti Shorkot',
 'Batkhela',
 'Behra',
 'Bhagtanwala',
 'Bhai Pheru',
 'Bhakkar',
 'Bhalwal',
 'Bhan Syedabad',
 'Bhiria',
 'Burewala',
 'Chak Jhumra',
 'Chakwal',
 'Charsadda',
 'Chashma',
 'Chattar',
 'Chawinda',
 'Chenab Nagar',
 'Chicha watni',
 'Chiniot',
 'Chishtian',
 'Choa Saidan Shah',
 'Chota Lahore',
 'Chowk Azam',
 'Chowk Munda',
 'Chowk Qureshi',
 'Dadu',
 'Dadyal',
 'Daharki',
 'Dahranwala',
 'Daluwali',
 'Darya Khan',
 'Daska',
 'Daur',
 'Depalpur',
 'Dera Allah Yar',
 'Dera Ghazi Khan',
 'Dera Ismail Khan',
 'Dera Murad Jamali',
 'Dhoria',
 'Dijkot',
 'Dina',
 'Dinga',
 'Donga Bonga',
 'Dunyapur',
 'Faisalabad',
 'Faqir Wali',
 'Farooka',
 'Fateh Pur',
 'Fatima Fertilizer Company',
 'Fazilpur',
 'Feroza',
 'Fort Abbas',
 'Gaggoo Mandi',
 'Gambat',
 'Garhi Dupatta',
 'Ghotki',
 'Gojra',
 'Gondal',
 'Goth Machi',
 'Guddu',
 'Gujar Khan',
 'Gujranwala',
 'Gujrat',
 'Hadli',
 'Hafizabad',
 'Hala',
 'Haripur',
 'Haroonabad',
 'Harrapa',
 'Hasan Abdal',
 'Hasil Pur',
 'Hatiyan',
 'Hattar',
 'Haveli Lakha',
 'Havelian',
 'Hazro',
 'Head Marralla',
 'Hujra Shah Muqeem',
 'Hyderabad',
 'In-Active',
 'Iqbal Nagar',
 'Isakhel',
 'Iskanderabad',
 'Islamabad',
 'Islamkot',
 'Jacobabad',
 'Jafarabad',
 'Jahanian',
 'Jalal Pur',
 'Jalal Pur Jattan',
 'Jalalpur Pir Wala',
 'Jamesabad',
 'Jampur',
 'Jamshoro',
 'Jaranwala',
 'Jatlan',
 'Jatoi',
 'Jauharabad',
 'Jehangira',
 'Jetha Bhutta',
 'Jhang',
 'Jhelum',
 'Jhuddo',
 'Jund',
 'Kabal',
 'Kabir Wala',
 'Kacha Khuh',
 'Kahror Pakka',
 'Kahuta',
 'Kala Bagh',
 'Kala Shah Kaku',
 'Kallar Kahar',
 'Kallar Syedan',
 'Kamaliya',
 'Kamar Mushani',
 'Kamir',
 'Kamoke',
 'Kamra',
 'Kandh Kot',
 'Kandiaro',
 'Karachi',
 'Kashmore',
 'Kassowal',
 'Kasur',
 'Katlang',
 'Khairabad',
 'Khairpur Mirs',
 'Khairpur Nathan Shah',
 'Khairpur Tamiwali',
 'Khalabat Township',
 'Khan Pur',
 'Khanewal',
 'Khanpur Mahar',
 'Kharian',
 'Kharota Syedan',
 'Khichi Wala',
 'Khurrianwala',
 'Khushab',
 'Kohat',
 'Kot Addu',
 'Kot Chutta',
 'Kot Ghulam Muhammad',
 'Kot Momin',
 'Kot Sabzal',
 'Kot Samaba',
 'Kotli',
 'Kotli Loharan',
 'Kotri',
 'Kunri',
 'Lahore',
 'Lalamusa',
 'Lalian',
 'Larkana',
 'Layyah',
 'Liaquat Pur',
 'Lodhran',
 'Lower Dir',
 'Luddan',
 'Machi Wal',
 'Mailsi',
 'Malakwal',
 'Malka Hans',
 'Mandi',
 'Mandi Bahauddin',
 'Mandi Madressa',
 'Mandrah',
 'Mandrah',
 'Mangla',
 'Mansehra',
 'Mardan',
 'Marot',
 'Matiari',
 'Mehmood Kot',
 'Merajke',
 'Mian Channu',
 'Mianwali',
 'Minchin Abad',
 'Mingora',
 'Mirpur Azad Kashmir',
 'Mirpur Khas',
 'Mirpur Mathelo',
 'Mirwah',
 'Mithi',
 'Mitro',
 'Moro',
 'Multan',
 'Murree',
 'Musafir Khana',
 'Muzaffar Garh',
 'Muzaffarabad',
 'Nankana Sahib',
 'Narowal',
 'Naudero',
 'Naushahro Feroze',
 'Nawabshah',
 'Noor Pur Thal',
 'Noor Shah',
 'Nowshera',
 'Nowshera Virkan',
 'Okara',
 'Okara Cantt',
 'Pahar Pur',
 'Paigah',
 'Pakpattan',
 'Panjeri',
 'Pano Akil',
 'Pasrur',
 'Pattoki',
 'Peshawar',
 'Petaro',
 'Phalia',
 'Phool Nagar',
 'Pind Dadan Khan',
 'Pindi Bhatiyan',
 'Piplan',
 'Pir Bala',
 'Pir Mahal',
 'Qaboola',
 'Qadirpur Rawan',
 'Qaim Pur',
 'Qasba Gujrat',
 'Qazi Ahmed',
 'Qila Didar Singh',
 'Quaidabad',
 'Quetta',
 'Rahim Yar Khan',
 'Rajan Pur',
 'Ranipur',
 'Rato Dero',
 'Rawalakot',
 'Rawalpindi',
 'Rawat',
 'Renala Khurd',
 'Risalpur',
 'Rohri',
 'Rustam',
 'Sadiqabad',
 'Sahiwal',
 'Saidan',
 'Saidu Sharif',
 'Sakhi Sarwar',
 'Sakrand',
 'Sambrial',
 'Samundri',
 'Sanawan',
 'Sanghar',
 'Sangla Hill',
 'Sanjar Pur',
 'Sarai Alamgir',
 'Sargodha',
 'Satiana',
 'Sehwan',
 'Shabqadar',
 'Shah Kot',
 'Shah Pur',
 'Shahdad Pur',
 'Shahdadkot',
 'Shahdara',
 'Shahpur Chakar',
 'Shakargarh',
 'Sheikhupura',
 'Shewa Adda',
 'Shikarpur',
 'Shorkot',
 'Shuja Abad',
 'Sialkot',
 'Sillanwali',
 'Sukkur',
 'Swabi',
 'Swat',
 'Talagang',
 'Tandlianwala',
 'Tando Adam',
 'Tando Allah Yar',
 'Tando Jam',
 'Tando Jan Mohammad',
 'Tando Muhammad Khan',
 'Taranda Muhammad Panah',
 'Taranda Saway Khan',
 'Tarbela',
 'Taunsa Sharif',
 'Taxila',
 'Tiba Sultan Pur',
 'Timergarah',
 'Toba Tek Singh',
 'Topi',
 'Toru',
 'Uch Sharif',
 'Ugoki',
 'Umerkot',
 'Vehari',
 'Wah Cantt',
 'Wan Bhachran',
 'Wazirabad',
 'Yazman',
 'Zaffarwal',
 'Zahir Pir']


class Gui(Tk):
    def __init__(self):
        super().__init__()
    def window_size(self,width=700,height=500):
        self.geometry(f'{width}x{height}')


def str_reverse(st):
    st = list(st)
    st.reverse()
    return ''.join(st)

def get_file(var):
    gui.file_get_dir = askopenfilename(filetype=(('Excel file', '*.xlsx'),('All files', '*.*')))
    var.set('')
    var.set(gui.file_get_dir)

def file_destination(var):
    file_set_position = askdirectory()
    var.set('')
    var.set(file_set_position)
def done():

    try:
        st = str_reverse(e1.get())
        index = st.index('/')
        file_xlsx = e1.get()[-index:]
        file_get_position = e1.get()[:-index]

        os.chdir(file_get_position)

        # a checker is available for correct format of file
        # an exception handler is for checking that the file is present or not
        if ".xlsx" not in file_xlsx:
            error.set('Incorrect file format or type, try again')
        else:
            error.set('')
            load_wb = load_workbook(file_xlsx)  # creat an object of Load_workbook and get data from an xl file
            sheet_list = list(load_wb.sheetnames)  # get list of number of sheets and store it in a variable
            load_wb_sheet = load_wb[sheet_list[0]]  # get the load workbook sheet and store it in a variable

            heading_list = ['Service Type ID', 'Pickup Address ID', 'Show Information on Air Waybill',
                            'Consignee City Name',
                            'Consignee Name', 'Consignee Address', 'Consignee Phone Number 1 (03000000000)',
                            'Consignee Phone Number 2 (03000000000)', 'Consignee Email Address', 'Order ID',
                            'Item Product Type ID', 'Item Description', 'Item Quantity', 'Item Insurance',
                            'Product Value',
                            'Replacement Item Product Type ID', 'Replacement Item Description',
                            'Replacement Item Quantity',
                            'Pickup Date (YYYY-MM-DD)', 'Special Instructions', 'Estimated Weight (kg)',
                            'Mode of Shipment ID',
                            'Same Day Timing ID', 'Collection Amount', 'Mode of Payment ID', 'Charges Mode ID']
            creat_wb = Workbook()  # creat an object of Workwook class
            creat_wb_sheet = creat_wb['Sheet']  # get the xl sheet and by name and store it in a variable
            num = 0

            for x in heading_list:  # this loop will initialize the creat_wb_sheet with items of heading list
                num += 1
                creat_wb_sheet.cell(row=1, column=num).value = x

            row = 1
            for x in range(1, load_wb_sheet.max_row + 1):  # this loop will run for max rows in load_wb_sheet times
                row += 1
                if load_wb_sheet[f'A{x}'].value is None or len(str(load_wb_sheet[f'A{x}'].value)) < 1:
                    row -=1
                    continue

                for y in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:  # number of columns in load_wb_sheet

                    if y == 'A':
                        creat_wb_sheet[f'J{row}'].value = load_wb_sheet[f'A{x}'].value

                    elif y == 'B':
                        creat_wb_sheet[f'A{row}'].value = service_type_id
                        creat_wb_sheet[f'B{row}'].value = pickup_address_id
                        creat_wb_sheet[f'C{row}'].value = 'no'
                        creat_wb_sheet[f'K{row}'].value = item_product_type_id

                    elif y == 'C':
                        creat_wb_sheet[f'E{row}'].value = load_wb_sheet[f'C{x}'].value

                    elif y == 'D':
                        creat_wb_sheet[f'L{row}'].value = load_wb_sheet[f'D{x}'].value
                        creat_wb_sheet[f'M{row}'].value = item_quantity
                        creat_wb_sheet[f'N{row}'].value = 'no'

                    elif y == 'E':
                        creat_wb_sheet[f'S{row}'].value = pickup_Date
                        creat_wb_sheet[f'T{row}'].value = special_instructions
                        creat_wb_sheet[f'U{row}'].value = estimated_weight

                    elif y == 'F':
                        creat_wb_sheet[f'V{row}'].value = mode_of_shipment_id['overnight']
                        creat_wb_sheet[f'X{row}'].value = load_wb_sheet[f'F{x}'].value

                    elif y == 'G':
                        creat_wb_sheet[f'Y{row}'].value = mode_of_payment_id
                        creat_wb_sheet[f'Z{row}'].value = charges_mode_id

                    elif y == 'H':
                        creat_wb_sheet[f'F{row}'].value = load_wb_sheet[f'H{x}'].value

                    elif y == 'I':
                        c = load_wb_sheet[f'I{x}'].value
                        if c.title() in list_of_cities:
                            creat_wb_sheet[f'D{row}'].value = c.title()
                        else:
                            pass

                    elif y == 'J':
                        number = str(load_wb_sheet[f'J{x}'].value)

                        if len(number) == 10 and number[0] == '3':
                            creat_wb_sheet[f'G{row}'].value = '0' + number

                        elif len(number) == 11 and number[0:2] == "03":
                            creat_wb_sheet[f'G{row}'].value = number
                        else:
                            pass

                    else:
                        pass

            load_wb.close()

            os.chdir(e2.get())
            creat_wb.save('Shipment Booking list.xlsx')

    except ZeroDivisionError as f:
        error.set(f)

gui = Gui()
gui.window_size()
gui.title('Idealancy COD Slip')
gui.config(bg= 'white')
Label(text='IDEALANCY',font='arial 20 bold',bg='white').pack(fill='x')

get_file_entry = StringVar()
f1 = Frame(bg='White').pack(side='left')
Label(f1,text='Get File',font='arial 13 bold',bg='White').pack(anchor='w',side='top')
Button(f1,text="Browse File",command=lambda:get_file(get_file_entry)).pack(anchor='w')
e1 = Entry(f1,textvariable=get_file_entry,border='3')
e1.pack(anchor='w',ipadx=150)

Label(text='__________________________________',bg='white').pack(anchor='w',pady=30)

file_destination_entry = StringVar()
f2 = Frame(bg='White',pady=30).pack(side='left')
Label(f2,text='File Destinaton',font='arial 13 bold',bg='White').pack(anchor='w',side='top')
Button(f2,text="Browse File",command=lambda:file_destination(file_destination_entry)).pack(anchor='w')
e2 = Entry(f2,textvariable=file_destination_entry,border='3')
e2.pack(anchor='w',ipadx=150)

error = StringVar()

Button(text='Done',command=done).pack(side='bottom',anchor='e',ipadx=30,ipady=30)

Label(textvariable=error,fg='red',bg='White',font='2').pack(side='top',anchor='w',pady=30)

gui.mainloop()

